from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook, load_workbook

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

import os
import sys

from calculations import calculate_cpk
from drawings import show_drawings_cpk, drawing_heatmap

FILENAME = "cpk_template.xlsx"


def check_none_values(value):
    """In case USL or LSl is emty or doesn't limited"""
    if value is None or str(value).strip() == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def select_excel_file(root):
    root.withdraw()
    
    initial_dir = os.path.expanduser("~")  # home folder
    
    file_path = filedialog.askopenfilename(
        title="Выберите файл данных для расчета",
        initialdir=initial_dir,
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("Excel 97-2003", "*.xls"),
            ("All files", "*.*")
        ]
    )
    
    if file_path:
        if os.path.exists(file_path):
            return file_path
        else:
            messagebox.showerror("Ошибка", "Файл не существует")
            root.deiconify()
            return None
    else:
        root.deiconify()
        return None
    

def upload_template_cpk(root):
    """Func to upload template for CPk"""
    file_path = select_excel_file(root)
    wb = load_workbook(file_path)
    ws = wb["Values"]

    collected_data = {
        'usl': check_none_values(ws['B2'].value),  # USL из ячейки B2
        'lsl': check_none_values(ws['B3'].value),  # LSL из ячейки B3
        'values': []       # измерения из столбца B
    }

    for row in range(5, 304):
        data_values = ws[f'B{row}'].value
        if data_values is not None:
            try:
                collected_data['values'].append(float(data_values))
            except (ValueError, TypeError):
                messagebox.showerror("Ошибка", "В данных есть значения, которые не удалось опредедлить как цифровые")


    results, values = calculate_cpk(collected_data)
    show_drawings_cpk(results, values)

    root.deiconify()

    
def upload_template_heatmap(root):
    """Func to upload template for HeatMap"""

    def collect_heatmap_data(tittle, col):
        for row in range(2, 304):
            data_values = ws[f'{col}{row}'].value
            if data_values is not None:
                try:
                    if tittle in collected_data:
                        collected_data[tittle].append(float(data_values))
                    else:
                        collected_data[tittle] = [float(data_values)]
                except (ValueError, TypeError):
                    messagebox.showerror("Ошибка", f"В данных есть значения, которые не удалось определить как цифру: строка {row}")


    file_path = select_excel_file(root)
    wb = load_workbook(file_path)
    ws = wb["Values"]

    collected_data = {}
    for tittle, col in zip((ws['C1'].value, ws['D1'].value, ws['E1'].value), ('C', 'D', 'E')):
        collect_heatmap_data(tittle, col)

    drawing_heatmap(collected_data)
    
    root.deiconify()




def download_template(root, status_label):
    """Upload template in xlsx to local hard drive"""
    wb = Workbook()
    ws = wb.active

    # wb styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E7E7E7", end_color="E7E7E7", fill_type="solid")
    usl_lsl = PatternFill(start_color="FFCFCF", end_color="FFCFCF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.title = "Values"
    ws['A2'] = "USL"
    ws['A3'] = "LSL"

    for row in range(2, 4):
        cell = ws.cell(row=row, column=1)
        cell.fill = usl_lsl
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment

    ws['B1'] = "CPk data"
    ws['C1'] = "Heat"
    ws['D1'] = "Map"
    ws['E1'] = "Values"
    for i in range(1, 5):
        cell = ws.cell(row=1, column=i+1)
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.font = header_font
        cell.fill = header_fill
    
    for row, i in zip(range(4, 304), range(1, 301)):
        ws[f'A{row}'] = f'value_{i}'
        cell = ws.cell(row=row, column=1)
        cell.font = header_font
        cell.fill = header_fill

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment

    try:
        wb.save(FILENAME)
        status_label.config(text="✅ Шаблон выгружен", foreground="green")

        if getattr(sys, 'frozen', False):
            current_dir = os.path.dirname(sys.executable)  # in case compiled exe
        else:
            current_dir = os.path.dirname(__file__) # in case script

        file_path = os.path.join(current_dir, FILENAME)

        os.startfile(file_path)
        
    except PermissionError:
        status_label.config(text=f"❌ Ошибка: Нет прав на запись или файл открыт в другой программе", foreground="red")

    except Exception as e:
        print(e)
        status_label.config(text=f"❌ Ошибка при сохранении", foreground="red")