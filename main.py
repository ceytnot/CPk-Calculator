# C:\Users\NewHorizon\Documents\PythonProjects\cpk_calculator\venv\Scripts\Activate.ps1
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

import matplotlib.pyplot as plt
import numpy as np

import os
import sys

import math

FILENAME = "cpk_template.xlsx"

def select_excel_file():
    root.withdraw()
    
    initial_dir = os.path.expanduser("~")  # home folder
    
    file_path = filedialog.askopenfilename(
        title="Выберите файл данных для расчета CPK",
        initialdir=initial_dir,
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("Excel 97-2003", "*.xls"),
            ("All files", "*.*")
        ]
    )
    
    if file_path:
        if os.path.exists(file_path):
            return file_path
        else:
            messagebox.showerror("Ошибка", "Файл не существует")
            root.deiconify()
            return None
    else:
        root.deiconify()
        return None


def download_template():
    wb = Workbook()
    ws = wb.active

    # wb styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E7E7E7", end_color="E7E7E7", fill_type="solid")
    usl_lsl = PatternFill(start_color="FFCFCF", end_color="FFCFCF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.title = "Values"
    ws['A2'] = "USL"
    ws['A3'] = "LSL"

    for row in range(2, 4):
        cell = ws.cell(row=row, column=1)
        cell.fill = usl_lsl
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment

    for col, i in zip(('B', 'C', 'D', 'E', 'F'), range(1, 6)):
        ws[f'{col}1'] = f"data_{i}"
        cell = ws.cell(row=1, column=i+1)
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.font = header_font
        cell.fill = header_fill
    
    for row, i in zip(range(4, 304), range(1, 301)):
        ws[f'A{row}'] = f'value_{i}'
        cell = ws.cell(row=row, column=1)
        cell.font = header_font
        cell.fill = header_fill

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment

    try:
        wb.save(FILENAME)
        status_label.config(text="✅ Шаблон выгружен", foreground="green")

        if getattr(sys, 'frozen', False):
            current_dir = os.path.dirname(sys.executable)  # in case compiled exe
        else:
            current_dir = os.path.dirname(__file__) # in case script

        file_path = os.path.join(current_dir, FILENAME)

        os.startfile(file_path)
        
    except PermissionError:
        status_label.config(text=f"❌ Ошибка: Нет прав на запись или файл открыт в другой программе", foreground="red")

    except Exception as e:
        print(e)
        status_label.config(text=f"❌ Ошибка при сохранении", foreground="red")



def normal_pdf(x, mean, std):
    """Probability Density Function (PDF) of Normal Distribution"""
    if std <= 0:
        return np.zeros_like(x)
    return (1.0 / (std * math.sqrt(2.0 * math.pi))) * np.exp(-((x - mean) ** 2) / (2.0 * std ** 2))

def show_drawings(results, values):
    """Counts parameters based on values and shows drawing in Matplotlib"""
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10))

    # Histogram
    n, bins, patches = ax1.hist(values, bins=10, edgecolor='black', alpha=0.6, 
                               color='skyblue', label='Данные')
    
    # Get parameters
    mean = results['mean']
    std = results['std_dev']
    usl = results['usl']
    lsl = results['lsl']
    
    # Target value
    target = results.get('target', (usl + lsl) / 2)
    
    # Curve of Probability Density Function (PDF)
    x_min = min(mean - 4 * std, lsl - 1, min(values) - 1)
    x_max = max(mean + 4 * std, usl + 1, max(values) + 1)
    x = np.linspace(x_min, x_max, 500)
    
    # SCALE the normal distribution curve to fit the histogram
    bin_width = bins[1] - bins[0]
    y_scaled = normal_pdf(x, mean, std) * len(values) * bin_width
    ax1.plot(x, y_scaled, 'r-', linewidth=2.5, label='Нормальное распределение')

    # get max_y
    max_y = max(y_scaled)

    # Central line of PDF
    ax1.axvline(mean, color='red', linestyle='-', linewidth=2, 
               label=f"Центр НР (μ): {mean:.2f}", alpha=0.7)
    
    # Target central line
    ax1.axvline(target, color='purple', linestyle='-.', linewidth=3, 
               label=f"Целевое значение (Target): {target:.2f}", alpha=0.8)
    
    # USL and LSL
    ax1.axvline(usl, color='green', linestyle='--', linewidth=1.5, 
               label=f"USL: {usl}", alpha=0.7)
    ax1.axvline(lsl, color='green', linestyle='--', linewidth=1.5, 
               label=f"LSL: {lsl}", alpha=0.7)

    # Levels for Sigma
    #colors = ['orange', 'gold', 'lightcoral']
    colors = ["#bfffb2", "#c3ff8a", "#f1ff72"]
  

    alphas = [0.3, 0.2, 0.1]
    
    for i in range(1, 4):
        # vertical lines for sigma
        ax1.axvline(mean + i*std, color='gray', linestyle=':', linewidth=1, alpha=0.5)
        ax1.axvline(mean - i*std, color='gray', linestyle=':', linewidth=1, alpha=0.5)
        
        # Fill areas for sigma
        if i == 1:
            ax1.axvspan(mean - std, mean + std, alpha=alphas[0], color=colors[0], 
                       label=f'±1σ (68.3%)')
            
            # Label 1σ inside area
            ax1.text(mean, max_y * 0.5, '1σ', fontsize=10, 
                   ha='center', va='center', color='black', fontweight='bold',
                   bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))
            
        elif i == 2:
            ax1.axvspan(mean - 2*std, mean - std, alpha=alphas[1], color=colors[1])
            ax1.axvspan(mean + std, mean + 2*std, alpha=alphas[1], color=colors[1])
            ax1.fill_between([], [], [], color=colors[1], alpha=alphas[1], 
                           label=f'±2σ (95.5%)')
            
            # Label 2σ inside area
            ax1.text(mean - 1.5*std, max_y * 0.3, '2σ', fontsize=9, 
                   ha='center', va='center', color='black', fontweight='bold',
                   bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))
            ax1.text(mean + 1.5*std, max_y * 0.3, '2σ', fontsize=9, 
                   ha='center', va='center', color='black', fontweight='bold',
                   bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))
            
        elif i == 3:
            ax1.axvspan(mean - 3*std, mean - 2*std, alpha=alphas[2], color=colors[2])
            ax1.axvspan(mean + 2*std, mean + 3*std, alpha=alphas[2], color=colors[2])
            ax1.fill_between([], [], [], color=colors[2], alpha=alphas[2], 
                           label=f'±3σ (99.7%)')
            
            # Label 3σ inside area
            ax1.text(mean - 2.5*std, max_y * 0.15, '3σ', fontsize=8, 
                   ha='center', va='center', color='black', fontweight='bold',
                   bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))
            ax1.text(mean + 2.5*std, max_y * 0.15, '3σ', fontsize=8, 
                   ha='center', va='center', color='black', fontweight='bold',
                   bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))

    # Vertical line - Center of PDF
    ax1.plot([mean, mean], [0, max_y], color='blue', linestyle=':', linewidth=1, alpha=0.5)

    # Defect areas
    defect_values = [v for v in values if v < lsl or v > usl]
    if defect_values:
        ax1.scatter(defect_values, [0] * len(defect_values), 
                  color='red', s=50, zorder=5, marker='v', 
                  label=f'Брак ({len(defect_values)} шт.)')

    # Labels and titles
    ax1.set_xlabel('Значения', fontsize=12)
    ax1.set_ylabel('Частота', fontsize=12)
    ax1.set_title('Нормальное распределение', 
                 fontsize=14, fontweight='bold')
    ax1.legend(loc='upper right', fontsize=9)
    ax1.grid(True, alpha=0.3)

    # SECOND GRAPH
    # count indices for horizontal axis
    x_indices = list(range(1, len(values) + 1))
    
    # linear graph
    ax2.plot(x_indices, values, 'b-', linewidth=2, marker='o', 
             markersize=4, label='Измерения')
    
    # Горизонтальные линии границ допуска
    ax2.axhline(y=usl, color='green', linestyle='--', linewidth=1.5, 
                label=f'USL: {usl}', alpha=0.7)
    ax2.axhline(y=lsl, color='green', linestyle='--', linewidth=1.5, 
                label=f'LSL: {lsl}', alpha=0.7)
    
    # Горизонтальная линия среднего
    ax2.axhline(y=mean, color='red', linestyle='-', linewidth=2, 
                label=f'Среднее: {mean:.2f}', alpha=0.7)
    
    # Горизонтальная линия целевого значения
    ax2.axhline(y=target, color='purple', linestyle='-.', linewidth=2, 
                label=f'Целевое значение: {target:.2f}', alpha=0.8)
    
    # Sigma levels (areas) colors = ["#bfffb2", "#d7ffb2", "#deffb2"]
    ax2.fill_between(x_indices, mean - std, mean + std, alpha=0.2, color=colors[0], 
                     label=f'±1σ ({mean-std:.2f} - {mean+std:.2f})')
    ax2.fill_between(x_indices, mean - 2*std, mean + 2*std, alpha=0.1, color=colors[1], 
                     label=f'±2σ ({mean-2*std:.2f} - {mean+2*std:.2f})')
    ax2.fill_between(x_indices, mean - 3*std, mean + 3*std, alpha=0.05, color=colors[2], 
                     label=f'±3σ ({mean-3*std:.2f} - {mean+3*std:.2f})')
    
    # Higlights for defects
    defect_indices = []
    defect_values_list = []
    for i, val in enumerate(values):
        if val < lsl or val > usl:
            defect_indices.append(x_indices[i])
            defect_values_list.append(val)
    
    if defect_indices:
        ax2.scatter(defect_indices, defect_values_list, color='red', s=80, 
                   zorder=5, marker='v', label=f'Брак ({len(defect_indices)} шт.)')
    
    # Properties of second graph
    ax2.set_xlabel('Номер измерения', fontsize=12)
    ax2.set_ylabel('Значение', fontsize=12)
    ax2.set_title('Линейный график измерений', fontsize=14, fontweight='bold')
    ax2.legend(loc='upper right', fontsize=9)
    ax2.grid(True, alpha=0.3)
    
    # Устанавливаем одинаковые пределы по Y для обоих графиков
    y_min = min(lsl - 0.5, min(values) - 0.5, mean - 4*std)
    y_max = max(usl + 0.5, max(values) + 0.5, mean + 4*std)
    ax2.set_ylim(y_min, y_max)

    # COMMON TEXT AND LABELS

    textstr = (f'Cp = {results["cp"]:.2f}\n'
               f'Cpk = {results["cpk"]:.2f}\n'
               #f'μ = {mean:.2f}\n'
               f'σ = {std:.2f}\n'
               #f'Target = {target:.2f}\n'
               f'Значений = {results["n"]}')

    
    # text block with statistic
    props = dict(boxstyle='round', facecolor='wheat', alpha=0.9)
    ax1.text(0.02, 0.98, textstr, transform=ax1.transAxes, fontsize=9,
            verticalalignment='top', bbox=props)

    plt.tight_layout()
    plt.show(block=False)

def calculate_cpk(data_values: dict) -> tuple[dict, list[float]]:
    """Returns results (cp, cpk, mean and so on...) and measurements"""
    
    usl = data_values['usl']
    lsl = data_values['lsl']
    measurements = data_values.get('values', [])
    
    if not measurements:
        messagebox.showerror("Ошибка", "Нет данных измерений")
    
    if usl is None or lsl is None:
        messagebox.showerror("Ошибка", "Не указаны USL или LSL")
    
    # Mean (среднее)
    n = len(measurements)
    mean = sum(measurements) / n
    
    # Standard deviation
    variance = sum((x - mean) ** 2 for x in measurements) / (n - 1)
    std_dev = math.sqrt(variance)
    
    # Indexes
    cp = (usl - lsl) / (6 * std_dev) if std_dev > 0 else float('inf')
    cpu = (usl - mean) / (3 * std_dev) if std_dev > 0 else float('inf')
    cpl = (mean - lsl) / (3 * std_dev) if std_dev > 0 else float('inf')
    cpk = min(cpu, cpl)

    # Target (Целевое значение)
    target = (usl + lsl) / 2 
    
    # Результаты
    results = {
        'n': n,
        'mean': mean,
        'std_dev': std_dev,
        'usl': usl,
        'lsl': lsl,
        'cp': cp,
        'cpk': cpk,
        'target': target
    }
    
    return results, measurements


def upload_template():

    file_path = select_excel_file()
    wb = load_workbook(file_path)
    ws = wb["Values"]

    collected_data = {
        'usl': float(ws['B2'].value),  # USL из ячейки B2
        'lsl': float(ws['B3'].value),  # LSL из ячейки B3
        'values': []       # измерения из столбца B
    }

    for row in range(4, 304):
        data_values = ws[f'B{row}'].value
        if data_values is not None:
            try:
                collected_data['values'].append(float(data_values))
            except (ValueError, TypeError):
                messagebox.showerror("Ошибка", "В данных есть значения, которые не удалось опредедлить как цифровые")



    results, values = calculate_cpk(collected_data)
    show_drawings(results, values)

    root.deiconify()



# ############## APP STARTS HERE #############################################################

root = tk.Tk()
root.title("CPk Calculator")
root.geometry("300x250")

# Styles
style_btn = ttk.Style()
style_btn.configure("main_btns.TButton", font=("Arial", 11), padding=10)


# Short Instruction
label_instruction = ttk.Label(root, text=
                  "1.Выгрузите шаблон в .xlsx\n" \
                  "2.Откройте его с помощью Excel\n" \
                  "3.Введите в шаблон имеющиеся данные\n" \
                  "4.Загрузите шаблон в программу для расчета\n   и постороения графиков")
label_instruction.pack(pady=10, padx=10)

# Frame for btns
button_frame = ttk.Frame(root)
button_frame.pack(pady=20)
status_label = ttk.Label(root, text="", font=("Arial", 10))
status_label.pack(pady=10)

# Buttons
submit_btn = ttk.Button(button_frame, text="Выгрузить\n Шаблон", style="main_btns.TButton", command=download_template)
submit_btn.pack(side='left', padx=5)

cancel_btn = ttk.Button(button_frame, text="Загрузить\n Шаблон", style="main_btns.TButton", command=upload_template)
cancel_btn.pack(side='left', padx=5)

root.mainloop()